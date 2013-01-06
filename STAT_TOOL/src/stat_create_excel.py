# -*- coding: utf-8 -*-


import datetime
from common import stat_util

from openpyxl import Workbook

INPUT_PATH = './input/david_data.csv'

SHEET_NAME_INDEX = 0
STAT_DATE_INDEX = 1
CT_INDEX = 2
ACTIVE_INDEX = 3

SHEET_INDEX = range(4,11)


#
def fix_row_data(rows):
    for row in rows:
        for i in range(0,len(row)):
            #print i
            row[i] = row[i].replace(')','').replace('(','')

def convert_rowdata_to_dict(datas):
    sheet_datas = {}
    for data in datas:
        print data
        sheet_name = data[SHEET_NAME_INDEX] + ' ' + data[ACTIVE_INDEX][0].upper() + data[ACTIVE_INDEX][1:]
        stat_date = data[STAT_DATE_INDEX]
        ct = data[CT_INDEX]
        if sheet_name not in sheet_datas:
            sheet_datas[sheet_name] = []
        appending_data = [stat_date + ',' + ct,data]
        #for debug
        if '2012-12-01' in sheet_name:
            print data
        #if appending_data not in sheet_datas[sheet_name]:
        has_include = False
        idx = 0
        for current_data in sheet_datas[sheet_name]:
            if appending_data[0] == current_data[0]:
                has_include = True
                break
            idx += 1
        if not has_include:
            sheet_datas[sheet_name].append(appending_data)
        else:
            sheet_datas[sheet_name][idx] = appending_data
        #print sheet_name,stat_date
        #print sheet_name
    return  sheet_datas

# create excel file in momery
def create_xlsx_from_dict(sheet_datas):
    wb = Workbook()
    SHEET_HEADER = [' ','Users','# Check-in','# Friends','# Active Friends','# KOL','# Posts in Feed','Sync SNS' ]
    keylist = sheet_datas.keys()
    keylist.sort()
    keylist.reverse()
    for key in keylist:
        sheet_name = key
        datas = sheet_datas.get(key)
        sorted(datas,cmp=lambda x, y:cmp(str(x[1][1]), str(y[1][1])))
        ws = wb.create_sheet(0, sheet_name)
        #for header
        for row in ws.range('A1:H1'):
            for i , cell in enumerate(row):
                cell.value = SHEET_HEADER[i]
        #for body
        row_idx = write_day_data(ws,datas)
        write_week_data(ws,datas,row_idx)
    return wb

def write_day_data(ws,datas):
    row_idx = 1
    #export_datas = {}
    for data in datas:
        data = data[1]
        sheet_name_date_str = data[SHEET_NAME_INDEX]
        stat_date_str = data[STAT_DATE_INDEX]
        sheet_date = datetime.datetime.strptime(sheet_name_date_str,'%Y-%m-%d')
        stat_date = datetime.datetime.strptime(stat_date_str,'%Y-%m-%d')
        values = []
        values.append('Day ' + str((stat_date-sheet_date).days + 1) +'_'+ data[CT_INDEX])
        for i in SHEET_INDEX:
            if i > SHEET_INDEX[0]:
                values.append('%.4f'  % ( int(data[i]) * 1.0 / int(data[SHEET_INDEX[0]])))
            else:
                values.append(data[i])
        #print values
        for col_idx , value in enumerate(values):
            cell = ws.cell(row = row_idx, column = col_idx)
            cell.value = value
        row_idx += 1
    return row_idx
    
def write_week_data(ws,datas,row_idx = 1):
    #week_num = 1
    week_datas = {}
    loop_counter = 0
    for data in datas:
        data = data[1]
        sheet_name_date_str = data[SHEET_NAME_INDEX]
        stat_date_str = data[STAT_DATE_INDEX]
        if stat_date_str == sheet_name_date_str:
            #print 'continue'
            continue
        sheet_date = datetime.datetime.strptime(sheet_name_date_str,'%Y-%m-%d') 
        stat_date = datetime.datetime.strptime(stat_date_str,'%Y-%m-%d') - datetime.timedelta(days = 1)
        week_num = str(int(((stat_date - sheet_date).days + 1 / 7.0) / 7 + 1)) +'_'+ data[CT_INDEX]
        if week_num not in week_datas:
            week_datas[week_num] = []
            
        week_datas[week_num].append(data)
    #print week_datas
    for key in sorted(week_datas.keys()):
        #week_num = 'Week' + str(key)
        #print week_num
        data = week_datas.get(key)
        new_data = [ 0 for x in range(0,11)]
        for i in range(0,len(data)):
            d = data[i]
            for x in range(0,4):
                new_data[x] = d[x]
            for j in SHEET_INDEX:
                """
                if j == SHEET_INDEX[0]:
                    new_data[j] = int(d[j])
                    continue
                """
                #print len(new_data), len(d)
                #new_data[j] += int(d[j])
                if j == SHEET_INDEX[0] or j == SHEET_INDEX[2]  or j == SHEET_INDEX[3] or j == SHEET_INDEX[4] or j == SHEET_INDEX[6]:#Post in feed----注册day2-day8，7天之内post的平均数
                    if j == SHEET_INDEX[0] and new_data[j] < int(d[j]):
                        new_data[j] = int(d[j]) #friends----day8 的好友数
                    else:
                        new_data[j] = int(d[j])
                else:
                    new_data[j] += int(d[j])
        week_datas[key] = new_data
    #print week_datas
    for key in sorted(week_datas.keys()):
        week_num = 'Week ' + str(key)
        data = week_datas.get(key)
        values = []
        values.append(week_num)
        for i in SHEET_INDEX:
            if i == SHEET_INDEX[0]:
                values.append(int(data[i]))
            elif i == SHEET_INDEX[2]  or i == SHEET_INDEX[3] or i == SHEET_INDEX[4] or i == SHEET_INDEX[6]:
                values.append(int(data[i]))
            elif i > SHEET_INDEX[0]:
                values.append('%.4f'  % ( int(data[i]) * 1.0 / int(data[SHEET_INDEX[0]]) ))
            else:
                #values.append(data[i] / 7)
                values.append(data[i])
        #print values
        for col_idx , value in enumerate(values):
            cell = ws.cell(row = row_idx, column = col_idx)
            cell.value = value
        row_idx += 1
        
if __name__ == '__main__':
    datas = stat_util.get_list_from_csv(INPUT_PATH)
    fix_row_data(datas)
    sheet_datas = convert_rowdata_to_dict(datas)
    #sheet_datas = stat_util.sort_dict_by_key(sheet_datas)
    wb = create_xlsx_from_dict(sheet_datas)
    wb.save('Data.xlsx')