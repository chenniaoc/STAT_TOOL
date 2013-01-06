# -*- coding: utf-8 -*-


import datetime
from common import stat_util

from common.csv_writer import CommonCsvWriter

from openpyxl import Workbook

from log.pflog import ProfilingLog,LengthLog,ClientLog,ImgLog

INPUT_PATH = './input/david_data.csv'

SHEET_NAME_INDEX = 0
STAT_DATE_INDEX = 1
CT_INDEX = 2
ACTIVE_INDEX = 3

SHEET_INDEX = range(4,11)


#
def fix_row_data(rows):
    for row in rows:
        for i in range(0,len(row)):
            #print i
            row[i] = row[i].replace(')','').replace('(','')

def convert_rowdata_to_dict(datas):
    sheet_datas = {}
    for data in datas:
        sheet_name = data[SHEET_NAME_INDEX] + ' ' + data[ACTIVE_INDEX][0].upper() + data[ACTIVE_INDEX][1:]
        stat_date = data[STAT_DATE_INDEX]
        ct = data[CT_INDEX]
        if sheet_name not in sheet_datas:
            sheet_datas[sheet_name] = []
        sheet_datas[sheet_name].append([stat_date + ',' + ct,data])
        #print sheet_name,stat_date
        #print sheet_name
    return  sheet_datas

# create excel file in momery
def create_xlsx_from_dict(sheet_datas):
    wb = Workbook()
    SHEET_HEADER = [' ','Users','# Check-in','# Friends','# Active Friends','# KOL','# Posts in Feed','Sync SNS' ]
    keylist = sheet_datas.keys()
    keylist.sort()
    keylist.reverse()
    for key in keylist:
        sheet_name = key
        datas = sheet_datas.get(key)
        sorted(datas,cmp=lambda x, y:cmp(str(x[1][1]), str(y[1][1])))
        ws = wb.create_sheet(0, sheet_name)
        #for header
        for row in ws.range('A1:H1'):
            for i , cell in enumerate(row):
                cell.value = SHEET_HEADER[i]
        #for body
        row_idx = write_day_data(ws,datas)
        write_week_data(ws,datas,row_idx)
    return wb

def write_day_data(ws,datas):
    row_idx = 1
    #export_datas = {}
    for data in datas:
        data = data[1]
        sheet_name_date_str = data[SHEET_NAME_INDEX]
        stat_date_str = data[STAT_DATE_INDEX]
        sheet_date = datetime.datetime.strptime(sheet_name_date_str,'%Y-%m-%d')
        stat_date = datetime.datetime.strptime(stat_date_str,'%Y-%m-%d')
        values = []
        values.append('Day ' + str((stat_date-sheet_date).days + 1) +'_'+ data[CT_INDEX])
        for i in SHEET_INDEX:
            if i > SHEET_INDEX[0]:
                values.append('%.4f'  % ( int(data[i]) * 1.0 / int(data[SHEET_INDEX[0]])))
            else:
                values.append(data[i])
        #print values
        for col_idx , value in enumerate(values):
            cell = ws.cell(row = row_idx, column = col_idx)
            cell.value = value
        row_idx += 1
    return row_idx
    
def write_week_data(ws,datas,row_idx = 1):
    #week_num = 1
    week_datas = {}
    for data in datas:
        data = data[1]
        sheet_name_date_str = data[SHEET_NAME_INDEX]
        stat_date_str = data[STAT_DATE_INDEX]
        sheet_date = datetime.datetime.strptime(sheet_name_date_str,'%Y-%m-%d')
        stat_date = datetime.datetime.strptime(stat_date_str,'%Y-%m-%d')
        week_num = str(int(((stat_date - sheet_date).days + 1 / 7.0) / 7 + 1)) +'_'+ data[CT_INDEX]
        if week_num not in week_datas:
            week_datas[week_num] = []
        week_datas[week_num].append(data)
    #print week_datas
    for key in sorted(week_datas.keys()):
        #week_num = 'Week' + str(key)
        #print week_num
        data = week_datas.get(key)
        new_data = [ 0 for x in range(0,11)]
        for i in range(0,len(data)):
            d = data[i]
            for x in range(0,4):
                new_data[x] = d[x]
            for j in SHEET_INDEX:
                """
                if j == SHEET_INDEX[0]:
                    new_data[j] = int(d[j])
                    continue
                """
                #print len(new_data), len(d)
                new_data[j] += int(d[j])
        week_datas[key] = new_data
    #print week_datas
    for key in sorted(week_datas.keys()):
        week_num = 'Week ' + str(key)
        data = week_datas.get(key)
        values = []
        values.append(week_num)
        for i in SHEET_INDEX:
            if i > SHEET_INDEX[0]:
                values.append('%.4f'  % ( int(data[i]) * 1.0 / int(data[SHEET_INDEX[0]]) ))
            else:
                values.append(data[i] / 7)
        #print values
        for col_idx , value in enumerate(values):
            cell = ws.cell(row = row_idx, column = col_idx)
            cell.value = value
        row_idx += 1

def export_csv(type = 'api'):
        #ProfilingLog
        #ClientLog
        test_pro = './input/looper_logs/api_server/profiling.log'
        test_client = './input/looper_logs/client/all_client.log'
        test_length = './input/looper_logs/api_length/length.log'
        test_img = './input/looper_logs/img_server/profiling.log'
        export_type = 'csv'
        
        pdatas = ProfilingLog.read_rows(test_pro)
        cdatas = ClientLog.read_rows(test_client)
        result = ProfilingLog.compare_to_client_log(pdatas,cdatas[1])
        lengthlog = LengthLog.read_rows(test_length)
        LengthLog.append_length_to_data(lengthlog,result)
        
        
        imglog = ImgLog.read_rows(test_img)
        img_result = ImgLog.compare_to_client_log(imglog, cdatas[1])
        if export_type == 'json':
            result_data = {'items':[]}
        else:
            result_data = []
        
        csw = CommonCsvWriter('compare_result_apiserver_with_client.csv')
        csw.write_header(['aa'])
        
        csw_img = CommonCsvWriter('compare_result_apiserver_with_client_img.csv')
        csw_img.write_header(['aa'])
        # apiserver loop
        for key in result:
            data = {}
            
            if type == 'api':
                data['req_id'] = key
                #row.append(key)
                log = result.get(key)
                req_id = log['client_log'][ClientLog.REQ_ID]
                req_type = log['client_log'][ClientLog.TYPE]
                is_wifi = u'yes' if log['client_log'][ClientLog.IS_WIFI] else 'no'
                client_spend_time = log['client_log'][ClientLog.RES_TIME]
                #print (len(log['profiling_log']))
                #print log['profiling_log']
                for profiling_log in log['profiling_log']:
                    row = []
                    req_api = profiling_log[ProfilingLog.REQ_API]
                    api_server_received_time = profiling_log[ProfilingLog.REQ_TIME]
                    api_server_req_ip = profiling_log[ProfilingLog.REQ_IP]
                    api_server_spd_time = profiling_log[ProfilingLog.SPD_TIME]
                    api_server_req_uid = profiling_log[ProfilingLog.RED_UID]
                    #print profiling_log 
                    api_server_version = profiling_log[ProfilingLog.REQ_VER]
                    api_server_response_length = log['length_log'][LengthLog.LENGTH]
                    row.append(req_id)
                    row.append(req_type)
                    row.append(is_wifi)
                    row.append(client_spend_time)
                    row.append(req_api)
                    row.append(api_server_received_time)
                    row.append(api_server_req_ip)
                    row.append(api_server_spd_time)
                    row.append(api_server_req_uid)
                    row.append(api_server_version)
                    row.append(api_server_response_length)
                    csw.write_onerow(row)
            """
            imglogs = img_result.get(key)
            if not imglogs:
                continue
            req_id = imglogs['client_log'][ClientLog.REQ_ID]
            req_type = imglogs['client_log'][ClientLog.TYPE]
            is_wifi = u'yes' if imglogs['client_log'][ClientLog.IS_WIFI] else 'no'
            client_spend_time = log['client_log'][ClientLog.RES_TIME]
            for img_log in imglogs['img_log']:
                img_spend_time = img_log[ImgLog.SPD_TIME]
                img_length = img_log[ImgLog.LENGTH]
                row.append(req_id)
                row.append(req_type)
                row.append(is_wifi)
                row.append(client_spend_time)
                row.append(img_spend_time)
                row.append(img_length)
                csw_img.write_onerow(row)


            
            """
            """
            if export_type == 'json':
                data['profiling_log'] = log['profiling_log']
                data['client_log'] = log['client_log']
                data['length_log'] = log['length_log']
                
                result_data['items'].append(data)
            else:
                data['length'] = log['length_log'][LengthLog.LENGTH]
                result_data.append(data)
            """
        for key in img_result:
            imglogs = img_result.get(key)
            if not imglogs:
                continue
            req_id = imglogs['client_log'][ClientLog.REQ_ID]
            if not req_id:
                continue
            req_type = imglogs['client_log'][ClientLog.TYPE]
            is_wifi = u'yes' if int(imglogs['client_log'][ClientLog.IS_WIFI]) else 'no'
            client_spend_time = imglogs['client_log'][ClientLog.RES_TIME]
            for img_log in imglogs['img_log']:
                img_type = img_log[ImgLog.TYPE]
                img_spend_time = img_log[ImgLog.SPD_TIME]
                img_length = img_log[ImgLog.LENGTH]
                
                row = []
                row.append(req_id)
                row.append(req_type)
                row.append(is_wifi)
                row.append(client_spend_time)
                row.append(img_spend_time)
                row.append(img_type)
                row.append(img_length)
                csw_img.write_onerow(row)
        csw.end_write()
        csw_img.end_write()
        return result_data
    

if __name__ == '__main__':
    export_csv()